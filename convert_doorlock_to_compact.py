#!/usr/bin/env python3
"""将门锁售后 Excel 文件转换为紧凑 JSON 格式，供 Web 分析系统使用。"""
import json, sys, os, time, re
from collections import Counter
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("需要 openpyxl，正在安装...")
    os.system(f"{sys.executable} -m pip install openpyxl")
    import openpyxl


# ── 文件路径：自动发现桌面最新 Excel 文件，或通过命令行参数指定 ──
import glob as _glob

def _find_latest(desktop, pattern):
    """返回桌面匹配 pattern 的最新文件"""
    files = _glob.glob(os.path.join(desktop, pattern))
    if not files:
        return None
    return max(files, key=os.path.getmtime)

DESKTOP = os.path.join(os.environ.get("USERPROFILE", ""), "Desktop")

# 检测输入模式：单文件合并 vs 双文件分离
MERGED_FILE = None
if len(sys.argv) > 1 and sys.argv[1].lower().endswith(('.xlsx', '.xls')):
    # 探测是否为合并文件（含"售后明细"sheet）
    _wb = openpyxl.load_workbook(sys.argv[1], data_only=True, read_only=True)
    if "售后明细" in _wb.sheetnames:
        MERGED_FILE = sys.argv[1]
    _wb.close()

if MERGED_FILE:
    AFTERSALE_FILE = SALES_FILE = MERGED_FILE
    INPUT_MODE = "merged"
    print(f"[合并模式] 读取: {MERGED_FILE}")
else:
    AFTERSALE_FILE = (
        sys.argv[1] if len(sys.argv) > 1 and not MERGED_FILE and not sys.argv[1].lower().endswith(('.xlsx', '.xls')) else
        _find_latest(DESKTOP, "亚马逊门锁售后工单*.xlsx")
    )
    SALES_FILE = (
        _find_latest(DESKTOP, "亚马逊门锁销量统计*.xlsx")
    )
    INPUT_MODE = "split"
    if not AFTERSALE_FILE or not os.path.exists(AFTERSALE_FILE):
        print("[X] 未找到售后工单文件，请在桌面放置 '亚马逊门锁售后工单*.xlsx' 或拖入 xlsx 到 一键更新.bat")
        sys.exit(1)
    if not SALES_FILE or not os.path.exists(SALES_FILE):
        print("[X] 未找到销量统计文件，请在桌面放置 '亚马逊门锁销量统计*.xlsx'")
        sys.exit(1)

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "data")
COMPACT_FILE = os.path.join(OUTPUT_DIR, "after-sale-data-compact.json")
VERSION_FILE = os.path.join(OUTPUT_DIR, "version.json")


# ── 责任方分类规则 ──
def classify_responsibility(reason, buyer_note=""):
    """基于售后原因和买家备注分类责任方"""
    reason_lower = (reason or "").lower()
    note_lower = (buyer_note or "").lower()

    # 品控 - 产品质量/缺陷问题
    qc_keywords = [
        "缺陷", "defective", "质量不可接受", "quality unacceptable",
        "缺少零件", "missing parts", "与描述不符", "not as described",
        "调包", "switcheroo", "尺寸过大", "尺寸过小", "too large", "too small",
        "poor fit", "不合适", "图物不符", "颜色不对", "颜色不匹配",
        "不如预期", "质量差", "太小", "太大", "wrong item", "发错",
    ]
    for kw in qc_keywords:
        if kw in reason_lower or kw in note_lower:
            return "品控"

    # 仓库 - 亚马逊仓库/发货问题
    wh_keywords = [
        "物流中心损坏", "damaged by fc", "fc",
    ]
    for kw in wh_keywords:
        if kw in reason_lower:
            return "仓库"

    # 物流 - 运输/投递问题
    log_keywords = [
        "无法投递", "undeliverable", "运输途中损坏", "damaged by carrier",
        "未在预计时间内送达", "missed estimated delivery", "never arriv",
        "未送达",
    ]
    for kw in log_keywords:
        if kw in reason_lower:
            return "物流"

    # 客户 - 主观原因
    cust_keywords = [
        "不想要", "unwanted", "无意购买", "订购了错误", "ordered wrong",
        "误购", "misordered", "未提供原因", "no reason given",
        "找到更优惠", "better price", "未经授权", "unauthorized",
        "改变主意", "不需要", "不再需要", "不喜欢", "没有必要",
        "购买太多", "订了太多",
    ]
    for kw in cust_keywords:
        if kw in reason_lower or kw in note_lower:
            return "客户"

    # 采购 - 供应商/货源问题
    pur_keywords = [
        "停产", "下架", "不可售", "晚发货", "延期",
    ]
    for kw in pur_keywords:
        if kw in reason_lower:
            return "采购"

    # 运营
    ops_keywords = ["折扣", "discount", "tp邀评"]
    for kw in ops_keywords:
        if kw in reason_lower or kw in note_lower:
            return "运营"

    return "客户"  # 默认归类为客户


# ── 读取 Excel ──
def read_excel(filepath, sheet_name=None):
    print(f"读取: {filepath}" + (f" [sheet={sheet_name}]" if sheet_name else ""))
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(row)
    wb.close()
    print(f"  → {len(rows)} 行数据, 表头: {headers}")
    return headers, rows


# ── 主转换 ──
def main():
    t_start = time.time()
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # 1. 读取售后数据
    ar_sheet = "售后明细" if INPUT_MODE == "merged" else None
    ah, ar_rows = read_excel(AFTERSALE_FILE, ar_sheet)
    # 列映射 - 兼容两种格式
    # 完整格式: 订单号、店铺、国家、SKU、售后数量、售后类型、售后时间、订购时间、售后原因、买家备注
    # 精简格式: SKU、售后数量、售后类型、售后时间、订购时间、售后原因、买家备注
    col_map = {h: i for i, h in enumerate(ah)}
    is_compact = "订单号" not in col_map  # 精简格式无订单号列

    after_sale_records = []
    type_counter = Counter()
    reason_counter = Counter()
    resp_counter = Counter()
    yearly = {}

    print("转换售后数据...")
    for row in ar_rows:
        tid = str(row[col_map["订单号"]]).strip() if not is_compact and row[col_map["订单号"]] else ""
        sku_val = str(row[col_map["SKU"]]).strip() if row[col_map["SKU"]] else ""
        if not sku_val or not sku_val.startswith("MS"):
            continue

        rq_val = row[col_map["售后数量"]]
        try:
            rq_val = int(rq_val)
        except (ValueError, TypeError):
            rq_val = 0

        ct_val = row[col_map["售后时间"]]
        if isinstance(ct_val, datetime):
            ct_str = ct_val.strftime("%Y-%m-%d %H:%M:%S")
        else:
            ct_str = str(ct_val).strip() if ct_val else ""

        # 提取年份
        year = int(ct_str[:4]) if ct_str and ct_str[:4].isdigit() else 0
        if year < 2020:
            # try from order time
            ot_val = row[col_map["订购时间"]]
            if isinstance(ot_val, datetime):
                year = ot_val.year
            elif ot_val:
                ot_str = str(ot_val).strip()
                if ot_str[:4].isdigit():
                    year = int(ot_str[:4])

        reason = str(row[col_map["售后原因"]]).strip() if row[col_map["售后原因"]] else ""
        buyer_note = str(row[col_map["买家备注"]]).strip() if row[col_map["买家备注"]] else ""
        after_type = str(row[col_map["售后类型"]]).strip() if row[col_map["售后类型"]] else ""

        # 责任方分类
        resp = classify_responsibility(reason, buyer_note)

        # 统计
        type_counter[after_type] += 1
        reason_counter[reason] += 1
        resp_counter[resp] += 1
        if year > 0:
            if year not in yearly:
                yearly[year] = {"after_sale_count": 0}
            yearly[year]["after_sale_count"] += 1

        after_sale_records.append({
            "tid": tid,
            "t": after_type,
            "sku": sku_val,
            "r": reason,
            "rq": rq_val,
            "tg": buyer_note,
            "ct": ct_str,
            "resp": resp,
            "y": year,
        })

    print(f"  售后记录: {len(after_sale_records)}")
    print(f"  售后类型: {dict(type_counter)}")
    print(f"  责任方: {dict(resp_counter)}")

    # 2. 读取销量数据
    sr_sheet = "销量明细" if INPUT_MODE == "merged" else None
    sh, sr_rows = read_excel(SALES_FILE, sr_sheet)
    # 列映射: 0:时间 1:SKU 2:销量 3:订单量
    s_col_map = {h: i for i, h in enumerate(sh)}

    sales_records = []
    total_sales = 0
    total_orders = 0

    print("转换销量数据...")
    for row in sr_rows:
        sku_val = str(row[s_col_map["SKU"]]).strip() if row[s_col_map["SKU"]] else ""
        if not sku_val or not sku_val.startswith("MS"):
            continue

        date_val = row[s_col_map["时间"]]
        if isinstance(date_val, datetime):
            d_str = date_val.strftime("%Y-%m-%d")
        else:
            d_str = str(date_val).strip()[:10] if date_val else ""
        year = int(d_str[:4]) if d_str[:4].isdigit() else 0

        sq_val = row[s_col_map["销量"]]
        oq_val = row[s_col_map["订单量"]]
        try:
            sq_val = int(sq_val)
            oq_val = int(oq_val)
        except (ValueError, TypeError):
            sq_val = oq_val = 0

        total_sales += sq_val
        total_orders += oq_val

        if year > 0:
            if year not in yearly:
                yearly[year] = {"after_sale_count": 0}
            if "sales_count" not in yearly[year]:
                yearly[year]["sales_count"] = 0
            yearly[year]["sales_count"] += sq_val

        sales_records.append({
            "d": d_str,
            "sku": sku_val,
            "sq": sq_val,
            "oq": oq_val,
            "y": year,
        })

    print(f"  销量记录: {len(sales_records)}")
    print(f"  总销量: {total_sales}, 总订单量: {total_orders}")

    # 3. 构建 compact JSON
    # 统计总 SKU 数
    all_sku = set(r["sku"] for r in after_sale_records) | set(r["sku"] for r in sales_records)

    compact = {
        "m": {
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "total_after_sale": len(after_sale_records),
            "total_sales": total_sales,
            "total_orders": total_orders,
            "unique_sku": len(all_sku),
            "type": "doorlock",
        },
        "ys": yearly,
        "st": {
            "after_sale_types": dict(type_counter),
            "reasons": dict(reason_counter.most_common()),
            "responsibilities": dict(resp_counter),
        },
        "ar": after_sale_records,
        "sr": sales_records,
    }

    # 4. 写入文件
    print(f"\n写入紧凑 JSON: {COMPACT_FILE}")
    with open(COMPACT_FILE, "w", encoding="utf-8") as f:
        json.dump(compact, f, ensure_ascii=False, separators=(",", ":"))
    file_size = os.path.getsize(COMPACT_FILE)
    print(f"  文件大小: {file_size / 1024 / 1024:.2f} MB")

    # 5. 生成 version.json
    version_data = {
        "version": datetime.now().strftime("%Y-%m-%d-v") + "1",
        "generated_at": compact["m"]["generated_at"],
        "total_after_sale": len(after_sale_records),
        "total_sales": total_sales,
        "total_orders": total_orders,
    }
    with open(VERSION_FILE, "w", encoding="utf-8") as f:
        json.dump(version_data, f, ensure_ascii=False, indent=2)
    print(f"写入版本文件: {VERSION_FILE}")

    elapsed = time.time() - t_start
    print(f"\n✅ 转换完成，耗时 {elapsed:.1f}s")


if __name__ == "__main__":
    main()
